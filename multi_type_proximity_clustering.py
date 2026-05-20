"""
Multi-Type Proximity Clustering

This project demonstrates clustering using:
- Nominal data
- Binary data
- Ordinal named data
- Cardinal numerical data
- Mixed-type data

Distance metrics and clustering techniques are applied
depending on the data type.

Author: Maria Rodopoulou
"""

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import gower

from sklearn.preprocessing import LabelEncoder, MinMaxScaler
from sklearn.metrics import pairwise_distances
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage


# ======================================================
# LOAD DATA
# ======================================================

def load_sheet(file_path, sheet_name):
    return pd.read_excel(file_path, sheet_name=sheet_name)


# ======================================================
# DISTANCE FUNCTIONS
# ======================================================

def compute_nominal_distance(df):
    df_encoded = df.astype(str).apply(LabelEncoder().fit_transform)
    return pairwise_distances(df_encoded, metric="hamming")


def compute_binary_distance(df):
    df_bin = df.replace({"ΝΑΙ": 1, "ΟΧΙ": 0, "Ναι": 1, "Όχι": 0, "NAI": 1, "OXI": 0})
    df_bin = df_bin.astype(int)
    return pairwise_distances(df_bin.values.astype(bool), metric="jaccard")


def compute_ordinal_distance(df, mapping_dict):
    df_clean = df.astype(str).map(lambda x: x.strip().lower())

    mapping_dict_lower = {
        k.lower(): v for k, v in mapping_dict.items()
    }

    df_mapped = df_clean.replace(mapping_dict_lower)
    df_mapped = df_mapped.apply(pd.to_numeric, errors="coerce")
    df_mapped = df_mapped.fillna(df_mapped.median())

    df_scaled = MinMaxScaler().fit_transform(df_mapped)

    return pairwise_distances(df_scaled, metric="euclidean")


def compute_cardinal_distance(df):
    df_num = df.apply(pd.to_numeric, errors="coerce")
    df_num = df_num.fillna(df_num.median())
    return pairwise_distances(df_num, metric="euclidean")


def compute_mixed_distance(df):
    """
    Gower distance for mixed data.

    Fix for newer pandas versions:
    gower package may not understand pandas StringDtype,
    so we convert everything to numpy object array.
    """
    df_clean = df.copy()

    for col in df_clean.columns:
        numeric_col = pd.to_numeric(df_clean[col], errors="coerce")

        if numeric_col.notna().sum() > 0:
            df_clean[col] = numeric_col.fillna(numeric_col.median())
        else:
            df_clean[col] = df_clean[col].astype(str).fillna("missing")

    df_clean = df_clean.astype(object)

    return gower.gower_matrix(df_clean.to_numpy(dtype=object))


# ======================================================
# CLUSTERING + VISUALIZATION
# ======================================================

def cluster_and_plot(dist_matrix, prefix):
    model = KMeans(n_clusters=3, random_state=42, n_init=10)
    labels = model.fit_predict(dist_matrix)

    pca = PCA(n_components=2)
    reduced = pca.fit_transform(dist_matrix)

    plt.figure(figsize=(7, 5))
    sns.scatterplot(
        x=reduced[:, 0],
        y=reduced[:, 1],
        hue=labels,
        palette="tab10"
    )

    plt.title(f"{prefix} - Clustering")
    plt.xlabel("PCA Component 1")
    plt.ylabel("PCA Component 2")
    plt.legend(title="Cluster")
    plt.tight_layout()

    fig_path = f"{prefix}_plot.png"
    plt.savefig(fig_path, dpi=150)
    plt.close()

    return labels, fig_path


# ======================================================
# MAIN PIPELINE
# ======================================================

def run_multi_type_clustering(
    file_path,
    output_excel="multi_type_proximity_results.xlsx"
):
    ordinal_mapping = {
        "Λύκειο": 0.0,
        "Προπτυχιακό": 0.33,
        "Μεταπτυχιακό": 0.66,
        "Διδακτορικό": 1.0,
        "Καθόλου": 0.0,
        "Μέτρια": 0.5,
        "Ικανοποιητική": 1.0
    }

    distance_functions = {
        "Nominal": compute_nominal_distance,
        "Binary": compute_binary_distance,
        "Ordinal_Named": lambda df: compute_ordinal_distance(df, ordinal_mapping),
        "Cardinal": compute_cardinal_distance,
        "Mixed": compute_mixed_distance
    }

    generated_images = []

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        for sheet_name, distance_function in distance_functions.items():
            print(f"▶ Processing: {sheet_name}")

            df = load_sheet(file_path, sheet_name)

            # Remove ID/index column if it exists
            if df.shape[1] > 1:
                df_features = df.iloc[:, 1:]
            else:
                df_features = df.copy()

            distance_matrix = distance_function(df_features)
            labels, img_path = cluster_and_plot(distance_matrix, sheet_name)
            generated_images.append((sheet_name, img_path))

            distance_df = pd.DataFrame(distance_matrix)
            clusters_df = pd.DataFrame({
                "Observation": range(1, len(labels) + 1),
                "Cluster": labels
            })

            distance_df.to_excel(
                writer,
                sheet_name=f"{sheet_name}_dist",
                index=False
            )

            clusters_df.to_excel(
                writer,
                sheet_name=f"{sheet_name}_clusters",
                index=False
            )

    # Insert images into Excel
    workbook = load_workbook(output_excel)

    if "Visualizations" in workbook.sheetnames:
        worksheet = workbook["Visualizations"]
    else:
        worksheet = workbook.create_sheet("Visualizations")

    for i, (sheet_name, img_path) in enumerate(generated_images):
        if os.path.exists(img_path):
            img = ExcelImage(img_path)
            img.anchor = f"A{1 + i * 20}"
            worksheet.add_image(img)

    workbook.save(output_excel)

    print(f"✅ Αποθηκεύτηκε στο: {output_excel}")


# ======================================================
# RUN SCRIPT
# ======================================================

if __name__ == "__main__":
    input_file = r"C:\Users\maria\Desktop\Clustering.Proximity.Software\multi_type_proximity_examples.xlsx"

    run_multi_type_clustering(
        input_file,
        output_excel=r"C:\Users\maria\Desktop\Clustering.Proximity.Software\multi_type_proximity_results.xlsx"
    )